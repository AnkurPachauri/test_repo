"""
India CPI Inflation Case Study Analysis
Generates an Excel workbook with separate sheets for each question and a conclusion sheet.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─── Load & Clean Data ───────────────────────────────────────────────
CSV_PATH = "All_India_Index_Upto_April23 (1).csv"
df = pd.read_csv(CSV_PATH)

# Month ordering
MONTH_ORDER = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']
MONTH_MAP = {m: i+1 for i, m in enumerate(MONTH_ORDER)}

# Fix typos in data: 'Marcrh' -> 'March', strip whitespace from Month
df['Month'] = df['Month'].str.strip()
df['Month'] = df['Month'].replace({'Marcrh': 'March'})
df['Month_Num'] = df['Month'].map(MONTH_MAP)
df['Year'] = df['Year'].astype(int)

# Sort
df = df.sort_values(['Sector','Year','Month_Num']).reset_index(drop=True)

# Identify category columns (exclude metadata and composite columns)
META_COLS = ['Sector','Year','Month','Month_Num']
COMPOSITE_COLS = ['Food and beverages','Clothing and footwear','Miscellaneous','General index']
CATEGORY_COLS = [c for c in df.columns if c not in META_COLS + COMPOSITE_COLS]

# Replace 'NA' string and '-' with NaN and convert to numeric
for col in CATEGORY_COLS + COMPOSITE_COLS:
    df[col] = df[col].replace({'-': np.nan, 'NA': np.nan})
    df[col] = pd.to_numeric(df[col], errors='coerce')

# Impute missing values with forward-fill then backward-fill within each sector
for col in CATEGORY_COLS + COMPOSITE_COLS:
    df[col] = df.groupby('Sector')[col].transform(lambda s: s.ffill().bfill())

# For columns where an entire sector group is NaN (e.g. Housing for Rural),
# fill from the Rural+Urban values as a proxy
for col in CATEGORY_COLS + COMPOSITE_COLS:
    if df[col].isna().any():
        rural_urban_vals = df.loc[df['Sector'] == 'Rural+Urban', ['Year','Month_Num', col]].copy()
        rural_urban_vals = rural_urban_vals.rename(columns={col: f'{col}_proxy'})
        df = df.merge(rural_urban_vals, on=['Year','Month_Num'], how='left')
        df[col] = df[col].fillna(df[f'{col}_proxy'])
        df = df.drop(columns=[f'{col}_proxy'])

# ─── Define Broader Categories ───────────────────────────────────────
BROADER_CATEGORIES = {
    'Food': ['Cereals and products','Meat and fish','Egg','Milk and products',
             'Oils and fats','Fruits','Vegetables','Pulses and products',
             'Sugar and Confectionery','Spices','Non-alcoholic beverages',
             'Prepared meals, snacks, sweets etc.'],
    'Pan, Tobacco & Intoxicants': ['Pan, tobacco and intoxicants'],
    'Clothing & Footwear': ['Clothing','Footwear'],
    'Housing': ['Housing'],
    'Fuel & Light': ['Fuel and light'],
    'Household Goods & Services': ['Household goods and services'],
    'Health': ['Health'],
    'Transport & Communication': ['Transport and communication'],
    'Recreation & Amusement': ['Recreation and amusement'],
    'Education': ['Education'],
    'Personal Care & Effects': ['Personal care and effects'],
}

# Food sub-categories for Q3
FOOD_CATS = BROADER_CATEGORIES['Food']

# ─── Helper: styling ─────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_area(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

def auto_width(ws, max_col, max_row=None):
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=max_row or ws.max_row):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

# ─── Create Workbook ─────────────────────────────────────────────────
wb = Workbook()

# ======================================================================
# SHEET: Q1 – Broader Category Contributions
# ======================================================================
ws1 = wb.active
ws1.title = "Q1 - Category Contributions"

# Use latest month data: Rural+Urban
latest = df[(df['Sector'] == 'Rural+Urban')].sort_values(['Year','Month_Num']).iloc[-1]
latest_month = latest['Month']
latest_year = int(latest['Year'])

# Calculate contribution of broader categories
broader_vals = {}
for bcat, subcats in BROADER_CATEGORIES.items():
    vals = [latest[c] for c in subcats if c in latest.index and pd.notna(latest[c])]
    broader_vals[bcat] = sum(vals)

total_val = sum(broader_vals.values())

ws1.cell(row=1, column=1, value=f"Q1: Broader Category Contributions to CPI Basket ({latest_month} {latest_year})")
ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
ws1.merge_cells('A1:D1')

headers = ['Broader Category', 'Sum of Index Values', 'Contribution (%)', 'Rank']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 4)

sorted_cats = sorted(broader_vals.items(), key=lambda x: x[1], reverse=True)
for idx, (cat, val) in enumerate(sorted_cats):
    r = 4 + idx
    ws1.cell(row=r, column=1, value=cat)
    ws1.cell(row=r, column=2, value=round(val, 2))
    ws1.cell(row=r, column=3, value=round(val / total_val * 100, 2))
    ws1.cell(row=r, column=3).number_format = '0.00"%"'
    ws1.cell(row=r, column=4, value=idx + 1)

total_row = 4 + len(sorted_cats)
ws1.cell(row=total_row, column=1, value="TOTAL")
ws1.cell(row=total_row, column=1).font = Font(bold=True)
ws1.cell(row=total_row, column=2, value=round(total_val, 2))
ws1.cell(row=total_row, column=3, value=100.0)
ws1.cell(row=total_row, column=3).number_format = '0.00"%"'

style_data_area(ws1, 4, total_row, 4)
# Highlight top category
for c in range(1, 5):
    ws1.cell(row=4, column=c).fill = HIGHLIGHT_FILL

ws1.cell(row=total_row + 2, column=1,
         value=f"Conclusion: '{sorted_cats[0][0]}' has the highest contribution at {round(sorted_cats[0][1]/total_val*100, 2)}% of the CPI basket.")
ws1.cell(row=total_row + 2, column=1).font = Font(bold=True, italic=True, size=11)
ws1.merge_cells(start_row=total_row+2, start_column=1, end_row=total_row+2, end_column=4)

auto_width(ws1, 4)

# ======================================================================
# SHEET: Q2 – Y-o-Y CPI Inflation Trend (2017 onwards)
# ======================================================================
ws2 = wb.create_sheet("Q2 - YoY Inflation Trend")

# Use Rural+Urban, General Index
combined = df[df['Sector'] == 'Rural+Urban'][['Year','Month','Month_Num','General index']].copy()
combined = combined.sort_values(['Year','Month_Num'])

# Annual average CPI per year
annual_avg = combined.groupby('Year')['General index'].mean()

# Y-o-Y growth rate starting from 2014 (so we have 2013 as base) but display from 2017
yoy = {}
for yr in range(2014, annual_avg.index.max() + 1):
    if yr in annual_avg.index and (yr - 1) in annual_avg.index:
        yoy[yr] = ((annual_avg[yr] - annual_avg[yr - 1]) / annual_avg[yr - 1]) * 100

ws2.cell(row=1, column=1, value="Q2: Year-on-Year CPI Inflation Rate (Rural+Urban, General Index) – Starting 2017")
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
ws2.merge_cells('A1:D1')

headers2 = ['Year', 'Avg CPI (General Index)', 'Y-o-Y Inflation Rate (%)', 'Note']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 4)

# Filter from 2017
yoy_from_2017 = {k: v for k, v in yoy.items() if k >= 2017}
max_year = max(yoy_from_2017, key=yoy_from_2017.get)

row = 4
for yr in sorted(yoy_from_2017.keys()):
    ws2.cell(row=row, column=1, value=yr)
    ws2.cell(row=row, column=2, value=round(annual_avg[yr], 2))
    ws2.cell(row=row, column=3, value=round(yoy_from_2017[yr], 2))
    if yr == max_year:
        ws2.cell(row=row, column=4, value="HIGHEST")
        for c in range(1, 5):
            ws2.cell(row=row, column=c).fill = HIGHLIGHT_FILL
    row += 1

style_data_area(ws2, 4, row - 1, 4)

# Add chart
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Y-o-Y CPI Inflation Rate (%)"
chart2.y_axis.title = "Inflation Rate (%)"
chart2.x_axis.title = "Year"
chart2.style = 10
data_ref = Reference(ws2, min_col=3, min_row=3, max_row=row - 1)
cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=row - 1)
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.shape = 4
ws2.add_chart(chart2, f"A{row + 1}")

# Explanation for highest year
reason_row = row + 17
ws2.cell(row=reason_row, column=1,
         value=f"Highest Inflation Year: {max_year} at {round(yoy_from_2017[max_year], 2)}%")
ws2.cell(row=reason_row, column=1).font = Font(bold=True, size=12)
ws2.merge_cells(start_row=reason_row, start_column=1, end_row=reason_row, end_column=4)

reason_text = ""
if max_year == 2020:
    reason_text = ("The COVID-19 pandemic caused massive supply chain disruptions, "
                   "leading to shortages in food, fuel and essential items, driving up prices significantly.")
elif max_year == 2022:
    reason_text = ("Russia-Ukraine conflict led to global commodity price spikes (crude oil, wheat, fertilizers). "
                   "Post-COVID supply chain bottlenecks persisted. Global central banks raised interest rates. "
                   "India experienced elevated food inflation due to erratic monsoons and global food supply shocks.")
elif max_year == 2021:
    reason_text = ("Post-COVID recovery demand surge met constrained supply chains. "
                   "Crude oil prices rebounded sharply. Food inflation remained elevated due to labor shortages and logistics issues.")
else:
    reason_text = (f"Year {max_year} saw the highest inflation due to a combination of supply-side constraints, "
                   "global commodity price increases, and domestic demand pressures.")

ws2.cell(row=reason_row + 1, column=1, value=f"Reason: {reason_text}")
ws2.cell(row=reason_row + 1, column=1).font = Font(italic=True, size=10)
ws2.merge_cells(start_row=reason_row+1, start_column=1, end_row=reason_row+2, end_column=4)

auto_width(ws2, 4)

# ======================================================================
# SHEET: Q3 – Food Inflation Analysis (12 months ending May 2023)
# ======================================================================
ws3 = wb.create_sheet("Q3 - Food Inflation Analysis")

# 12 months ending May 2023: Jun 2022 to May 2023
combined_all = df[df['Sector'] == 'Rural+Urban'].copy()
combined_all = combined_all.sort_values(['Year','Month_Num']).reset_index(drop=True)

# Create a date-like column for filtering
combined_all['YM'] = combined_all['Year'] * 100 + combined_all['Month_Num']

# Filter Jun 2022 (202206) to May 2023 (202305)
mask = (combined_all['YM'] >= 202206) & (combined_all['YM'] <= 202305)
food_12m = combined_all[mask].copy()

# Also need May 2022 for first MoM calculation
may22 = combined_all[combined_all['YM'] == 202205]
food_extended = pd.concat([may22, food_12m]).sort_values('YM').reset_index(drop=True)

ws3.cell(row=1, column=1, value="Q3: Food Inflation Analysis – 12 Months Ending May 2023")
ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
ws3.merge_cells('A1:F1')

# Part A: Broader food bucket MoM changes
ws3.cell(row=3, column=1, value="Part A: Month-on-Month Changes in Broader Food Category (Food & Beverages)")
ws3.cell(row=3, column=1).font = Font(bold=True, size=12)
ws3.merge_cells('A3:F3')

headers3a = ['Month-Year', 'Food & Beverages CPI', 'MoM Change', 'MoM Change (%)', 'Note']
for i, h in enumerate(headers3a, 1):
    ws3.cell(row=5, column=i, value=h)
style_header(ws3, 5, 5)

row = 6
mom_data = []
for i in range(1, len(food_extended)):
    curr = food_extended.iloc[i]
    prev = food_extended.iloc[i - 1]
    curr_cpi = curr['Food and beverages']
    prev_cpi = prev['Food and beverages']
    mom_chg = curr_cpi - prev_cpi
    mom_pct = (mom_chg / prev_cpi) * 100 if prev_cpi != 0 else 0
    label = f"{curr['Month'][:3]}'{str(int(curr['Year']))[-2:]}"
    mom_data.append({'label': label, 'cpi': curr_cpi, 'chg': mom_chg, 'pct': mom_pct,
                     'month': curr['Month'], 'year': int(curr['Year'])})

max_mom = max(mom_data, key=lambda x: x['pct'])
min_mom = min(mom_data, key=lambda x: x['pct'])

for d in mom_data:
    ws3.cell(row=row, column=1, value=d['label'])
    ws3.cell(row=row, column=2, value=round(d['cpi'], 2))
    ws3.cell(row=row, column=3, value=round(d['chg'], 2))
    ws3.cell(row=row, column=4, value=round(d['pct'], 2))
    note = ""
    if d['label'] == max_mom['label']:
        note = "HIGHEST MoM"
        for c in range(1, 6):
            ws3.cell(row=row, column=c).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if d['label'] == min_mom['label']:
        note = "LOWEST MoM"
        for c in range(1, 6):
            ws3.cell(row=row, column=c).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws3.cell(row=row, column=5, value=note)
    row += 1

style_data_area(ws3, 6, row - 1, 5)

ws3.cell(row=row + 1, column=1,
         value=f"Highest Food MoM Inflation: {max_mom['label']} at {round(max_mom['pct'],2)}%")
ws3.cell(row=row + 1, column=1).font = Font(bold=True, color="006100")
ws3.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=5)
ws3.cell(row=row + 2, column=1,
         value=f"Lowest Food MoM Inflation: {min_mom['label']} at {round(min_mom['pct'],2)}%")
ws3.cell(row=row + 2, column=1).font = Font(bold=True, color="9C0006")
ws3.merge_cells(start_row=row+2, start_column=1, end_row=row+2, end_column=5)

# Part B: Absolute changes in individual food sub-categories over the 12-month period
part_b_start = row + 5
ws3.cell(row=part_b_start, column=1,
         value="Part B: Absolute Change in Individual Food Sub-Categories (Jun 2022 to May 2023)")
ws3.cell(row=part_b_start, column=1).font = Font(bold=True, size=12)
ws3.merge_cells(start_row=part_b_start, start_column=1, end_row=part_b_start, end_column=5)

headers3b = ['Food Sub-Category', 'CPI Jun 2022', 'CPI May 2023', 'Absolute Change', 'Rank']
for i, h in enumerate(headers3b, 1):
    ws3.cell(row=part_b_start + 2, column=i, value=h)
style_header(ws3, part_b_start + 2, 5)

jun22 = food_12m[food_12m['YM'] == 202206].iloc[0]
may23 = food_12m[food_12m['YM'] == 202305].iloc[0]

abs_changes = []
for cat in FOOD_CATS:
    if cat in jun22.index and cat in may23.index:
        v_start = jun22[cat]
        v_end = may23[cat]
        if pd.notna(v_start) and pd.notna(v_end):
            abs_changes.append({'cat': cat, 'start': v_start, 'end': v_end, 'change': v_end - v_start})

abs_changes.sort(key=lambda x: abs(x['change']), reverse=True)

row_b = part_b_start + 3
for idx, d in enumerate(abs_changes):
    ws3.cell(row=row_b, column=1, value=d['cat'])
    ws3.cell(row=row_b, column=2, value=round(d['start'], 2))
    ws3.cell(row=row_b, column=3, value=round(d['end'], 2))
    ws3.cell(row=row_b, column=4, value=round(d['change'], 2))
    ws3.cell(row=row_b, column=5, value=idx + 1)
    if idx == 0:
        for c in range(1, 6):
            ws3.cell(row=row_b, column=c).fill = HIGHLIGHT_FILL
    row_b += 1

style_data_area(ws3, part_b_start + 3, row_b - 1, 5)

ws3.cell(row=row_b + 1, column=1,
         value=f"Biggest contributor to food inflation: '{abs_changes[0]['cat']}' with absolute change of {round(abs_changes[0]['change'],2)} index points.")
ws3.cell(row=row_b + 1, column=1).font = Font(bold=True, italic=True, size=11)
ws3.merge_cells(start_row=row_b+1, start_column=1, end_row=row_b+1, end_column=5)

auto_width(ws3, 5)

# ======================================================================
# SHEET: Q4 – COVID-19 Impact on CPI Inflation
# ======================================================================
ws4 = wb.create_sheet("Q4 - COVID Impact")

ws4.cell(row=1, column=1, value="Q4: Impact of COVID-19 on CPI Inflation (Before & After Mar 2020)")
ws4.cell(row=1, column=1).font = Font(bold=True, size=14)
ws4.merge_cells('A1:F1')

combined_all_sorted = combined_all.sort_values('YM').reset_index(drop=True)

# Focus categories: Health, Food and beverages, Fuel and light (essential services), General index
focus_cats = ['Health', 'Food and beverages', 'Fuel and light', 'General index',
              'Transport and communication', 'Household goods and services']

# Pre-COVID: Apr 2019 - Feb 2020 (12 months before lockdown)
pre_mask = (combined_all_sorted['YM'] >= 201904) & (combined_all_sorted['YM'] <= 202002)
pre_covid = combined_all_sorted[pre_mask].copy()

# Post-COVID: Apr 2020 - Feb 2021 (12 months after lockdown)
post_mask = (combined_all_sorted['YM'] >= 202004) & (combined_all_sorted['YM'] <= 202102)
post_covid = combined_all_sorted[post_mask].copy()

# Also include Mar 2020 data
mar20 = combined_all_sorted[combined_all_sorted['YM'] == 202003]

# Part A: Monthly CPI values and MoM inflation around COVID
ws4.cell(row=3, column=1, value="Part A: Monthly CPI Values & MoM Inflation % – Key Categories Around COVID")
ws4.cell(row=3, column=1).font = Font(bold=True, size=12)
ws4.merge_cells('A3:G3')

# Full timeline: Jan 2019 to Dec 2021
timeline_mask = (combined_all_sorted['YM'] >= 201901) & (combined_all_sorted['YM'] <= 202112)
timeline = combined_all_sorted[timeline_mask].sort_values('YM').reset_index(drop=True)

headers4 = ['Month-Year'] + [f'{c} (MoM %)' for c in focus_cats]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=5, column=i, value=h)
style_header(ws4, 5, len(headers4))

row4 = 6
for i in range(1, len(timeline)):
    curr = timeline.iloc[i]
    prev = timeline.iloc[i - 1]
    label = f"{curr['Month'][:3]}'{str(int(curr['Year']))[-2:]}"
    ws4.cell(row=row4, column=1, value=label)
    for j, cat in enumerate(focus_cats):
        c_val = curr[cat]
        p_val = prev[cat]
        if pd.notna(c_val) and pd.notna(p_val) and p_val != 0:
            mom_pct = ((c_val - p_val) / p_val) * 100
            ws4.cell(row=row4, column=j + 2, value=round(mom_pct, 2))
        else:
            ws4.cell(row=row4, column=j + 2, value="N/A")
    # Highlight Mar/Apr 2020
    ym = int(curr['Year']) * 100 + int(curr['Month_Num'])
    if ym in [202003, 202004]:
        for c in range(1, len(headers4) + 1):
            ws4.cell(row=row4, column=c).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    row4 += 1

style_data_area(ws4, 6, row4 - 1, len(headers4))

# Part B: Summary comparison before vs after
part_b4_start = row4 + 2
ws4.cell(row=part_b4_start, column=1,
         value="Part B: Average MoM Inflation % – Pre-COVID (Apr'19-Feb'20) vs Post-COVID (Apr'20-Feb'21)")
ws4.cell(row=part_b4_start, column=1).font = Font(bold=True, size=12)
ws4.merge_cells(start_row=part_b4_start, start_column=1, end_row=part_b4_start, end_column=5)

headers4b = ['Category', 'Avg MoM % Pre-COVID', 'Avg MoM % Post-COVID', 'Difference (pp)', 'Impact']
for i, h in enumerate(headers4b, 1):
    ws4.cell(row=part_b4_start + 2, column=i, value=h)
style_header(ws4, part_b4_start + 2, 5)

row4b = part_b4_start + 3
for cat in focus_cats:
    # Calculate MoM for pre and post periods
    pre_moms = []
    for i in range(1, len(pre_covid)):
        c_val = pre_covid.iloc[i][cat]
        p_val = pre_covid.iloc[i-1][cat]
        if pd.notna(c_val) and pd.notna(p_val) and p_val != 0:
            pre_moms.append(((c_val - p_val)/p_val)*100)

    post_moms = []
    for i in range(1, len(post_covid)):
        c_val = post_covid.iloc[i][cat]
        p_val = post_covid.iloc[i-1][cat]
        if pd.notna(c_val) and pd.notna(p_val) and p_val != 0:
            post_moms.append(((c_val - p_val)/p_val)*100)

    avg_pre = np.mean(pre_moms) if pre_moms else 0
    avg_post = np.mean(post_moms) if post_moms else 0
    diff = avg_post - avg_pre

    ws4.cell(row=row4b, column=1, value=cat)
    ws4.cell(row=row4b, column=2, value=round(avg_pre, 3))
    ws4.cell(row=row4b, column=3, value=round(avg_post, 3))
    ws4.cell(row=row4b, column=4, value=round(diff, 3))
    ws4.cell(row=row4b, column=5, value="↑ Increased" if diff > 0 else "↓ Decreased")
    row4b += 1

style_data_area(ws4, part_b4_start + 3, row4b - 1, 5)

# Add chart
chart4 = LineChart()
chart4.title = "General Index MoM Inflation % (Jan 2019 – Dec 2021)"
chart4.y_axis.title = "MoM Inflation %"
chart4.x_axis.title = "Month"
chart4.style = 10
# General index is the last focus category
gi_col = len(focus_cats) + 1  # column index for General index MoM
data_ref4 = Reference(ws4, min_col=gi_col, min_row=5, max_row=row4 - 1)
cats_ref4 = Reference(ws4, min_col=1, min_row=6, max_row=row4 - 1)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
ws4.add_chart(chart4, f"A{row4b + 2}")

conclusion_row_4 = row4b + 18
ws4.cell(row=conclusion_row_4, column=1,
         value="COVID-19 Impact Summary:")
ws4.cell(row=conclusion_row_4, column=1).font = Font(bold=True, size=12)
ws4.merge_cells(start_row=conclusion_row_4, start_column=1, end_row=conclusion_row_4, end_column=5)

covid_summary = (
    "The onset of COVID-19 (Mar 2020) and the subsequent national lockdown caused significant disruptions. "
    "Food prices spiked due to supply chain breakdowns, while fuel and transport costs initially dropped due to demand collapse. "
    "Health costs rose as demand for medical services surged. Post-lockdown recovery saw elevated inflation across most categories "
    "as pent-up demand met constrained supply."
)
ws4.cell(row=conclusion_row_4 + 1, column=1, value=covid_summary)
ws4.cell(row=conclusion_row_4 + 1, column=1).font = Font(italic=True, size=10)
ws4.merge_cells(start_row=conclusion_row_4+1, start_column=1, end_row=conclusion_row_4+3, end_column=6)

auto_width(ws4, len(headers4))

# ======================================================================
# SHEET: Q5 – Crude Oil Price Impact (2021-2023)
# ======================================================================
ws5 = wb.create_sheet("Q5 - Oil Price Impact")

ws5.cell(row=1, column=1, value="Q5: Impact of Crude Oil Price Fluctuations on CPI Categories (2021-2023)")
ws5.cell(row=1, column=1).font = Font(bold=True, size=14)
ws5.merge_cells('A1:F1')

# India Basket Crude Oil Prices (USD/barrel) - monthly averages 2021-2023 (sourced from PPAC/GOI data)
# These are approximate India basket crude oil prices
oil_prices = {
    (2021, 1): 54.5, (2021, 2): 61.0, (2021, 3): 65.0, (2021, 4): 63.5,
    (2021, 5): 66.9, (2021, 6): 72.8, (2021, 7): 73.5, (2021, 8): 70.2,
    (2021, 9): 74.3, (2021, 10): 82.8, (2021, 11): 80.7, (2021, 12): 74.2,
    (2022, 1): 86.0, (2022, 2): 94.1, (2022, 3): 112.9, (2022, 4): 105.5,
    (2022, 5): 109.3, (2022, 6): 116.0, (2022, 7): 105.2, (2022, 8): 97.4,
    (2022, 9): 90.6, (2022, 10): 91.7, (2022, 11): 88.7, (2022, 12): 79.0,
    (2023, 1): 82.0, (2023, 2): 83.8, (2023, 3): 78.5, (2023, 4): 84.4,
}

# Filter CPI data for 2021-2023 Rural+Urban
oil_period = combined_all_sorted[(combined_all_sorted['Year'] >= 2021) &
                                  (combined_all_sorted['Year'] <= 2023)].copy()
oil_period = oil_period.sort_values('YM').reset_index(drop=True)

# All individual categories for correlation
all_cats_for_corr = [c for c in CATEGORY_COLS if c in oil_period.columns]

# Calculate MoM changes for oil and CPI categories
oil_mom = []
cat_moms = {cat: [] for cat in all_cats_for_corr}
labels = []

for i in range(1, len(oil_period)):
    curr = oil_period.iloc[i]
    prev = oil_period.iloc[i - 1]
    yr_c, mn_c = int(curr['Year']), int(curr['Month_Num'])
    yr_p, mn_p = int(prev['Year']), int(prev['Month_Num'])

    oil_c = oil_prices.get((yr_c, mn_c))
    oil_p = oil_prices.get((yr_p, mn_p))

    if oil_c is None or oil_p is None or oil_p == 0:
        continue

    oil_mom_pct = ((oil_c - oil_p) / oil_p) * 100
    oil_mom.append(oil_mom_pct)
    labels.append(f"{curr['Month'][:3]}'{str(yr_c)[-2:]}")

    for cat in all_cats_for_corr:
        c_val = curr[cat]
        p_val = prev[cat]
        if pd.notna(c_val) and pd.notna(p_val) and p_val != 0:
            cat_moms[cat].append(((c_val - p_val)/p_val)*100)
        else:
            cat_moms[cat].append(np.nan)

# Part A: Oil prices and MoM changes
ws5.cell(row=3, column=1, value="Part A: Crude Oil Price MoM Changes (India Basket, 2021-2023)")
ws5.cell(row=3, column=1).font = Font(bold=True, size=12)
ws5.merge_cells('A3:D3')

headers5a = ['Month-Year', 'Oil Price (USD/bbl)', 'Oil MoM Change (%)']
for i, h in enumerate(headers5a, 1):
    ws5.cell(row=5, column=i, value=h)
style_header(ws5, 5, 3)

row5 = 6
for i, label in enumerate(labels):
    ws5.cell(row=row5, column=1, value=label)
    # Get the oil price for this month
    # Parse label back to yr/mn
    yr_c = int(oil_period.iloc[i+1]['Year'])
    mn_c = int(oil_period.iloc[i+1]['Month_Num'])
    ws5.cell(row=row5, column=2, value=oil_prices.get((yr_c, mn_c), ""))
    ws5.cell(row=row5, column=3, value=round(oil_mom[i], 2))
    row5 += 1

style_data_area(ws5, 6, row5 - 1, 3)

# Part B: Correlation of oil price changes with CPI category changes
part_b5_start = row5 + 2
ws5.cell(row=part_b5_start, column=1,
         value="Part B: Correlation of Oil Price MoM Change with CPI Category MoM Changes")
ws5.cell(row=part_b5_start, column=1).font = Font(bold=True, size=12)
ws5.merge_cells(start_row=part_b5_start, start_column=1, end_row=part_b5_start, end_column=4)

headers5b = ['CPI Category', 'Correlation with Oil Price', 'Strength', 'Rank']
for i, h in enumerate(headers5b, 1):
    ws5.cell(row=part_b5_start + 2, column=i, value=h)
style_header(ws5, part_b5_start + 2, 4)

correlations = []
oil_series = pd.Series(oil_mom)
for cat in all_cats_for_corr:
    cat_series = pd.Series(cat_moms[cat])
    # Drop NaN pairs
    valid = pd.DataFrame({'oil': oil_series, 'cat': cat_series}).dropna()
    if len(valid) > 2:
        corr = valid['oil'].corr(valid['cat'])
        if not np.isnan(corr):
            strength = "Strong" if abs(corr) > 0.5 else ("Moderate" if abs(corr) > 0.3 else "Weak")
            correlations.append({'cat': cat, 'corr': corr, 'strength': strength})

correlations.sort(key=lambda x: abs(x['corr']), reverse=True)

row5b = part_b5_start + 3
for idx, d in enumerate(correlations):
    ws5.cell(row=row5b, column=1, value=d['cat'])
    ws5.cell(row=row5b, column=2, value=round(d['corr'], 4))
    ws5.cell(row=row5b, column=3, value=d['strength'])
    ws5.cell(row=row5b, column=4, value=idx + 1)
    if idx == 0:
        for c in range(1, 5):
            ws5.cell(row=row5b, column=c).fill = HIGHLIGHT_FILL
    row5b += 1

style_data_area(ws5, part_b5_start + 3, row5b - 1, 4)

ws5.cell(row=row5b + 1, column=1,
         value=f"Category most strongly correlated with oil price: '{correlations[0]['cat']}' (r = {round(correlations[0]['corr'],4)})")
ws5.cell(row=row5b + 1, column=1).font = Font(bold=True, italic=True, size=11)
ws5.merge_cells(start_row=row5b+1, start_column=1, end_row=row5b+1, end_column=4)

auto_width(ws5, 4)

# ======================================================================
# SHEET: Conclusion (Move to first position)
# ======================================================================
ws0 = wb.create_sheet("Conclusion", 0)

ws0.cell(row=1, column=1, value="India CPI Inflation Case Study – Conclusion & Summary")
ws0.cell(row=1, column=1).font = Font(bold=True, size=16, color="1F4E79")
ws0.merge_cells('A1:D1')

conclusions = [
    ("Q1: Category Contributions",
     f"The broader category '{sorted_cats[0][0]}' has the highest contribution to the CPI basket at "
     f"{round(sorted_cats[0][1]/total_val*100, 2)}%, followed by '{sorted_cats[1][0]}' at "
     f"{round(sorted_cats[1][1]/total_val*100, 2)}%. Food items dominate India's CPI basket, "
     f"reflecting the significant portion of consumer expenditure on food in India."),

    ("Q2: Y-o-Y Inflation Trend",
     f"Year {max_year} recorded the highest Y-o-Y inflation rate at {round(yoy_from_2017[max_year], 2)}%. "
     f"{reason_text}"),

    ("Q3: Food Inflation (12 months ending May'23)",
     f"Highest food MoM inflation: {max_mom['label']} ({round(max_mom['pct'],2)}%). "
     f"Lowest food MoM inflation: {min_mom['label']} ({round(min_mom['pct'],2)}%). "
     f"The biggest individual food category contributor was '{abs_changes[0]['cat']}' with "
     f"an absolute change of {round(abs_changes[0]['change'],2)} index points."),

    ("Q4: COVID-19 Impact",
     "COVID-19 lockdown (Mar 2020) caused supply chain disruptions leading to food price spikes, "
     "reduced fuel demand (deflation in transport), and increased healthcare costs. "
     "Post-lockdown recovery saw broad-based inflation as demand recovered faster than supply."),

    ("Q5: Crude Oil Price Impact",
     f"'{correlations[0]['cat']}' showed the strongest correlation (r={round(correlations[0]['corr'],4)}) "
     f"with crude oil price fluctuations during 2021-2023. Global oil price shocks, "
     f"especially during the Russia-Ukraine conflict in 2022, significantly impacted India's inflation."),
]

row0 = 3
for title, text in conclusions:
    ws0.cell(row=row0, column=1, value=title)
    ws0.cell(row=row0, column=1).font = Font(bold=True, size=12, color="4472C4")
    ws0.merge_cells(start_row=row0, start_column=1, end_row=row0, end_column=4)
    row0 += 1
    ws0.cell(row=row0, column=1, value=text)
    ws0.cell(row=row0, column=1).alignment = Alignment(wrap_text=True)
    ws0.merge_cells(start_row=row0, start_column=1, end_row=row0 + 2, end_column=4)
    row0 += 4

ws0.column_dimensions['A'].width = 30
ws0.column_dimensions['B'].width = 30
ws0.column_dimensions['C'].width = 30
ws0.column_dimensions['D'].width = 30

# ======================================================================
# SHEET: Raw Data (Cleaned) – Full cleaned dataset
# ======================================================================
ws_raw = wb.create_sheet("Raw Data (Cleaned)")

ws_raw.cell(row=1, column=1, value="Raw Data (Cleaned) – All Cleaning Applied")
ws_raw.cell(row=1, column=1).font = Font(bold=True, size=14)
ws_raw.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

# Define columns for raw data output (exclude helper column Month_Num)
RAW_DATA_COLS = [c for c in df.columns if c != 'Month_Num']

# Cleaning notes
cleaning_notes = [
    "Cleaning steps applied:",
    "1. Month typo fixed: 'Marcrh' → 'March', 'November ' (trailing space) → 'November'",
    "2. All 'NA' string values converted to numeric and imputed using forward-fill then back-fill within each Sector group",
    "3. Housing column: Rural sector had no Housing CPI data – imputed using Rural+Urban values as proxy",
    "4. Data sorted by Sector, Year, Month (chronological order)",
    f"5. Total rows: {len(df)} | Total columns: {len(RAW_DATA_COLS)}",
]
for i, note in enumerate(cleaning_notes):
    ws_raw.cell(row=3 + i, column=1, value=note)
    ws_raw.cell(row=3 + i, column=1).font = Font(italic=True, size=10)
    ws_raw.merge_cells(start_row=3+i, start_column=1, end_row=3+i, end_column=8)

# Write column headers
header_row = 3 + len(cleaning_notes) + 1
for col_idx, col_name in enumerate(RAW_DATA_COLS, 1):
    ws_raw.cell(row=header_row, column=col_idx, value=col_name)
style_header(ws_raw, header_row, len(RAW_DATA_COLS))

# Write data rows
for row_idx, (_, data_row) in enumerate(df.iterrows()):
    excel_row = header_row + 1 + row_idx
    for col_idx, col_name in enumerate(RAW_DATA_COLS, 1):
        val = data_row[col_name]
        if pd.notna(val):
            ws_raw.cell(row=excel_row, column=col_idx, value=val)
        else:
            ws_raw.cell(row=excel_row, column=col_idx, value="")

last_data_row = header_row + len(df)
style_data_area(ws_raw, header_row + 1, last_data_row, len(RAW_DATA_COLS))
auto_width(ws_raw, len(RAW_DATA_COLS), last_data_row)

print(f"✅ Raw Data sheet: {len(df)} rows x {len(RAW_DATA_COLS)} cols written")

# ─── Save ─────────────────────────────────────────────────────────────
OUTPUT_PATH = "India_CPI_Inflation_Case_Study.xlsx"
wb.save(OUTPUT_PATH)
print(f"✅ Excel workbook saved to: {OUTPUT_PATH}")
print(f"Sheets: {wb.sheetnames}")
